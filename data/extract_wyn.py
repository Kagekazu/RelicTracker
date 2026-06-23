#!/usr/bin/env python3
"""Extract Wyn's Relic Tracker xlsx into JSON for RelicTracker plugin."""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Wyn's Relic Tracker.xlsx"
OUT = ROOT / "extracted"

EXPANSION_SHEETS = [
    "ARR",
    "HW",
    "SB",
    "ShB",
    "EW",
    "DT",
    "DoHDoL",
]

# Column headers vary by sheet; we locate them on row 2.
MATERIAL_HEADERS = {"Material", "Per Weapon", "Per Tool", "Held", "Remaining", "Progress", "Kettle"}
CURRENCY_MARKERS = {
    "Poetics",
    "Company Seals",
    "Allied Seals",
    "Gil",
    "MGP",
    "Bicolor",
    "Purple",
    "Orange",
    "White",
    "Tomestones",
}

COLUMN_HEADER_LABELS = {
    "Step",
    "Material",
    "Held",
    "Remaining",
    "Progress",
    "Per Weapon",
    "Per Tool",
    "Per Set",
    "Kettle",
}


def to_optional_number(value) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text or text in COLUMN_HEADER_LABELS:
            return None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def is_subsection_header_row(material: str | None, step_b: str | None) -> bool:
    if material in COLUMN_HEADER_LABELS:
        return True
    return step_b == "Step" and material == "Material"


def norm(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def find_header_columns(ws, header_row: int = 2) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        val = norm(ws.cell(header_row, col).value)
        if val:
            headers[val] = col
    return headers


def first_job_column() -> int:
    return 3  # column C


def is_section_header(step: str | None, material: str | None) -> bool:
    return step == "Step" and material == "Material"


def tag_material_expansions(rows: list[dict], expansion_sheets: dict[str, dict]) -> None:
    if not rows or any(row.get("expansion") for row in rows):
        return

    doh_steps = {
        norm(entry.get("step"))
        for entry in expansion_sheets.get("DoHDoL", {}).get("materials", [])
        if norm(entry.get("step"))
    }
    header_indices = [i for i, row in enumerate(rows) if is_section_header(row.get("step"), row.get("material"))]
    bounds = [0, *[index + 1 for index in header_indices], len(rows)]
    bounds = sorted(set(bounds))
    section_expansions = ["ARR", "HW", "SB", "ShB", "EW", "DT", "DoHDoL"]

    for section_index, section_expansion in enumerate(section_expansions):
        if section_index + 1 >= len(bounds):
            break

        start, end = bounds[section_index], bounds[section_index + 1]
        for row in rows[start:end]:
            if is_section_header(row.get("step"), row.get("material")):
                continue

            step = row.get("step")
            if section_expansion == "DT" and step in doh_steps:
                row["expansion"] = "DoHDoL"
            else:
                row["expansion"] = section_expansion


def extract_materials(ws) -> list[dict]:
    rows: list[dict] = []
    current_step: str | None = None
    for r in range(3, (ws.max_row or 0) + 1):
        step = norm(ws.cell(r, 1).value)
        material = norm(ws.cell(r, 2).value)
        location = norm(ws.cell(r, 3).value)
        requirement = norm(ws.cell(r, 4).value)
        note = norm(ws.cell(r, 5).value)

        if step:
            current_step = step.replace("\n", " ").strip()

        if not any([material, location, requirement, note]):
            continue

        rows.append(
            {
                "step": current_step,
                "material": material,
                "location": location,
                "requirement": requirement,
                "note": note,
            }
        )
    return rows


def is_currency_row(material: str | None) -> bool:
    if not material:
        return False
    return any(marker in material for marker in CURRENCY_MARKERS)


def extract_expansion(ws, sheet_id: str) -> dict:
    headers = find_header_columns(ws)
    material_col = headers.get("Material")
    if not material_col:
        raise ValueError(f"{sheet_id}: Material column not found")

    per_unit_col = headers.get("Per Weapon") or headers.get("Per Tool")
    held_col = headers.get("Held")
    remaining_col = headers.get("Remaining")
    progress_col = headers.get("Progress")
    kettle_col = headers.get("Kettle")

    job_start = first_job_column()
    job_end = material_col - 1
    job_count = max(0, job_end - job_start + 1)

    job_names: list[str] = []
    for c in range(job_start, job_end + 1):
        job_names.append(norm(ws.cell(2, c).value) or f"Job {len(job_names) + 1}")

    steps: list[dict] = []
    materials: list[dict] = []
    currencies: list[dict] = []
    current_step: str | None = None
    material_subsection: str | None = None

    for r in range(3, (ws.max_row or 0) + 1):
        step_a = norm(ws.cell(r, 1).value)
        step_b = norm(ws.cell(r, 2).value)

        if step_a:
            current_step = step_a.replace("\n", " ").strip()
            material_subsection = None

        material = norm(ws.cell(r, material_col).value) if material_col else None
        if material and is_currency_row(material):
            currencies.append(
                {
                    "name": material,
                    "held": to_optional_number(ws.cell(r, held_col).value if held_col else None),
                    "remaining": to_optional_number(ws.cell(r, remaining_col).value if remaining_col else None),
                    "progress": to_optional_number(ws.cell(r, progress_col).value if progress_col else None),
                    "perUnit": to_optional_number(ws.cell(r, per_unit_col).value if per_unit_col else None),
                }
            )
            continue

        if step_b == "Steps Remaining":
            jobs_remaining = []
            for c in range(job_start, job_end + 1):
                jobs_remaining.append(ws.cell(r, c).value)
            steps.append(
                {
                    "kind": "stepsRemaining",
                    "step": current_step,
                    "label": step_b,
                    "jobsRemaining": jobs_remaining,
                }
            )
            continue

        if not step_b and not material:
            continue

        if is_subsection_header_row(material, step_b):
            continue

        job_flags = []
        for c in range(job_start, job_end + 1):
            val = ws.cell(r, c).value
            job_flags.append(bool(val) if isinstance(val, bool) else None)

        if material in {"Crafters", "Fisher", "Miner & Botanist"} and not any(
            flag is not None for flag in job_flags
        ):
            material_subsection = material

        if not any(flag is not None for flag in job_flags) and material_subsection:
            group_ranges = {
                "Crafters": range(0, 8),
                "Miner & Botanist": range(8, 10),
                "Fisher": range(10, 11),
            }
            job_range = group_ranges.get(material_subsection)
            if job_range is not None:
                job_flags = [
                    (index in job_range) if index < job_count else None
                    for index in range(job_count)
                ]

        entry = {
            "step": current_step,
            "label": step_b,
            "jobs": job_flags,
            "material": material,
            "held": to_optional_number(ws.cell(r, held_col).value if held_col else None),
            "remaining": to_optional_number(ws.cell(r, remaining_col).value if remaining_col else None),
            "progress": to_optional_number(ws.cell(r, progress_col).value if progress_col else None),
            "perUnit": to_optional_number(ws.cell(r, per_unit_col).value if per_unit_col else None),
            "kettle": to_optional_number(ws.cell(r, kettle_col).value if kettle_col else None),
        }

        if material:
            materials.append(entry)
        elif step_b:
            steps.append(entry)

    return {
        "id": sheet_id,
        "jobCount": job_count,
        "jobNames": job_names,
        "jobColumnStart": openpyxl.utils.get_column_letter(job_start),
        "materialColumn": openpyxl.utils.get_column_letter(material_col),
        "steps": steps,
        "materials": materials,
        "currencies": currencies,
    }


def extract_manifest(wb) -> dict:
    landing = wb["Landing"]
    version = norm(landing.cell(3, 5).value) or norm(landing.cell(4, 5).value)
    patch = norm(landing.cell(3, 2).value)
    return {
        "source": "Wyn's Relic Tracker",
        "sheetVersion": version,
        "patch": patch,
        "expansions": EXPANSION_SHEETS,
    }


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing workbook: {XLSX}")

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    OUT.mkdir(parents=True, exist_ok=True)

    manifest = extract_manifest(wb)
    (OUT / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    materials = extract_materials(wb["Materials"])
    expansions: dict[str, dict] = {}
    for sheet_id in EXPANSION_SHEETS:
        expansions[sheet_id] = extract_expansion(wb[sheet_id], sheet_id)

    tag_material_expansions(materials, expansions)
    (OUT / "materials.json").write_text(json.dumps(materials, indent=2), encoding="utf-8")

    (OUT / "expansions.json").write_text(json.dumps(expansions, indent=2), encoding="utf-8")

    notes_rows = []
    ws = wb["Notes"]
    for r in range(2, (ws.max_row or 0) + 1):
        row = [norm(ws.cell(r, c).value) for c in range(1, 6)]
        if any(row):
            notes_rows.append(row)
    (OUT / "notes.json").write_text(json.dumps(notes_rows, indent=2), encoding="utf-8")

    wb.close()

    summary = {
        "manifest": manifest,
        "materialReferenceRows": len(materials),
        "expansions": {k: {"materials": len(v["materials"]), "steps": len(v["steps"])} for k, v in expansions.items()},
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
